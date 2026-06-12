"""
Checkmarx Security Report — Email Draft Generator
===================================================
Drop in same folder:
  - Checkmarx XML report      (*.xml)
  - Vulnerability Tracker     (*.xlsx)   ← optional but recommended
  - SLA image                 (*.png)    ← optional

Run:  python generate_email_draft.py
Open generated HTML → Copy Email Body → Paste into Outlook.
"""

import xml.etree.ElementTree as ET
import os, glob, sys, base64
from datetime import date, datetime, timedelta
from collections import defaultdict

# ── Config ────────────────────────────────────────────────────────────────────
BASE_URL  = "https://blackduck.phibred.com:444/srm/projects"
SEV_ORDER = ["critical", "high", "medium", "low", "info"]
SECTION_MAP = {
    "component analysis": "SCA",
    "static analysis":    "SAST",
    "iac analysis":       "IaC",
}
REM_SLA    = {"critical": None, "high": 45, "medium": 90, "low": 180, "info": None}
TBL_HEADER = "#BDD7EE"
# ─────────────────────────────────────────────────────────────────────────────


# ── File finders ──────────────────────────────────────────────────────────────
def find_file(ext):
    d = os.path.dirname(os.path.abspath(__file__))
    files = glob.glob(os.path.join(d, f"*.{ext}"))
    if not files:
        return None
    files.sort(key=os.path.getmtime, reverse=True)
    print(f"Found .{ext} : {os.path.basename(files[0])}")
    return files[0]


def find_png():
    d = os.path.dirname(os.path.abspath(__file__))
    files = glob.glob(os.path.join(d, "*.png"))
    if not files:
        return None
    files.sort(key=os.path.getmtime, reverse=True)
    with open(files[0], "rb") as f:
        data = base64.b64encode(f.read()).decode()
    return f"data:image/png;base64,{data}"


# ── Date helpers ──────────────────────────────────────────────────────────────
def parse_date(s):
    if not s:
        return None
    if hasattr(s, 'date'):          # already a datetime from openpyxl
        return s.date() if hasattr(s, 'hour') else s
    s = str(s).strip().replace("Z", "+00:00").replace("z", "+00:00")
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def rem_date_from(first_date, severity):
    sev = severity.lower()
    if sev == "critical":
        return "N/A", False
    if sev == "info":
        return "When possible", False
    days = REM_SLA.get(sev)
    if days is None or first_date is None:
        return "N/A", False
    rd = first_date + timedelta(days=days)
    overdue = rd < date.today()
    return rd.strftime("%d-%b-%Y"), overdue


def clean_code(s):
    return s.replace("_", " ") if s else ""


# ── Load tracker xlsx ─────────────────────────────────────────────────────────
def load_tracker(xlsx_path):
    """
    Returns dict keyed by (vuln_name_lower, section_lower):
      { "first_identified": date, "severity": str, "remediation": "Active"/"Fixed" }
    Columns used (case-insensitive header match):
      Application Name, First Identified, Vulnerability Severity,
      Vulnerability Name, Remediation, Assessment Name
    """
    if xlsx_path is None:
        return {}

    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl not installed — tracker skipped. Run: pip install openpyxl")
        return {}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Find sheet named "Vulnerability Tracker" or first sheet
    sheet = None
    for name in wb.sheetnames:
        if "vulnerability tracker" in name.lower() or "tracker" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    # Map header → column index (case-insensitive)
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    def col(name):
        for i, h in enumerate(header):
            if name.lower() in h:
                return i
        return None

    ci_vuln     = col("vulnerability name")
    ci_first    = col("first identified")
    ci_sev      = col("vulnerability severity")
    ci_rem      = col("remediation")
    ci_assess   = col("assessment")   # "Assessment Name" or "Assessment Type"

    if ci_vuln is None:
        print("⚠️  Could not find 'Vulnerability Name' column in tracker — skipped.")
        return {}

    tracker = {}
    for row in rows[1:]:
        try:
            vuln_name = str(row[ci_vuln]).strip() if row[ci_vuln] else ""
            if not vuln_name or vuln_name.lower() == "none":
                continue
            first_id  = parse_date(row[ci_first])  if ci_first  is not None else None
            severity  = str(row[ci_sev]).strip()    if ci_sev    is not None else ""
            rem       = str(row[ci_rem]).strip()    if ci_rem    is not None else "Active"
            assess    = str(row[ci_assess]).strip() if ci_assess is not None else ""

            key = (vuln_name.lower(), assess.lower())
            # Keep earliest first_identified if duplicate keys
            if key not in tracker or (first_id and tracker[key]["first_identified"] and
                                       first_id < tracker[key]["first_identified"]):
                tracker[key] = {
                    "first_identified": first_id,
                    "severity":         severity,
                    "remediation":      rem,
                    "vuln_name":        vuln_name,
                    "assess":           assess,
                }
        except Exception:
            continue

    print(f"  Tracker rows loaded: {len(tracker)}")
    return tracker


def tracker_lookup(tracker, vuln_name, section):
    """Try exact match first, then section-agnostic."""
    key1 = (vuln_name.lower(), section.lower())
    key2 = (vuln_name.lower(), "")
    if key1 in tracker:
        return tracker[key1]
    # try any section
    for k, v in tracker.items():
        if k[0] == vuln_name.lower():
            return v
    return None


# ── Parse XML ─────────────────────────────────────────────────────────────────
def parse_report(xml_path, tracker):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    raw_date = root.attrib.get("date", "")
    try:
        scan_dt   = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
        scan_date = scan_dt.strftime("%d-%b-%Y")
        scan_date_raw = scan_dt.date()
    except Exception:
        scan_date     = raw_date[:10] if raw_date else "N/A"
        scan_date_raw = date.today()

    proj      = root.find("project")
    app_name  = proj.attrib.get("name", "N/A") if proj is not None else "N/A"
    proj_id   = proj.attrib.get("id", "")      if proj is not None else ""
    portal    = f"{BASE_URL}/{proj_id}"         if proj_id else BASE_URL

    author_raw  = root.attrib.get("author", "")
    author_name = ""
    if author_raw:
        local       = author_raw.split("@")[0]
        author_name = local.replace(".", " ").replace("_", " ").title()

    groups = defaultdict(lambda: {
        "instances": 0, "status": "New", "rem_date": "",
        "overdue": False, "first_seen": "", "vuln_name": "",
        "row_red": False,
    })

    # Track all vuln names seen in XML per section (for fixed table)
    xml_vulns = {"SAST": set(), "SCA": set(), "IaC": set()}

    for finding in root.iter("finding"):
        det = finding.attrib.get("detection-method", "").lower()
        sec = SECTION_MAP.get(det)
        if sec is None:
            continue

        severity   = finding.attrib.get("severity", "low").lower()
        first_seen = finding.attrib.get("first-seen", "")
        fs_date    = parse_date(first_seen) or scan_date_raw

        rule_el  = finding.find("rule")
        vuln_cat = clean_code(rule_el.attrib.get("name", "Unknown")) if rule_el is not None else "Unknown"
        results  = finding.findall(".//result")
        inst_count = len(results)

        # ── Get vuln_name per section type ────────────────────────────────────
        if sec == "SCA":
            first_result = results[0] if results else None
            vuln_name    = vuln_cat
            if first_result is not None:
                candidates = (first_result.findall(".//Additional-value") +
                              first_result.findall(".//additional-value"))
                for node in candidates:
                    if node.attrib.get("key", "").strip().lower() == "component identifier":
                        vuln_name = node.text.strip() if node.text else vuln_cat
                        break

            finding_id = finding.attrib.get("id", str(id(finding)))
            gkey = (sec, finding_id, vuln_cat, severity)

            # Cross-reference tracker
            t = tracker_lookup(tracker, vuln_name, sec)
            if t is None:
                # Not in tracker → New
                rem_str, overdue = rem_date_from(fs_date, severity)
                status   = "New"
                row_red  = False
            elif t["remediation"].strip().lower() == "fixed":
                # Was fixed but reappeared → New, row red
                rem_str, overdue = rem_date_from(fs_date, severity)
                status   = "New"
                row_red  = True
            else:
                # Active in tracker → use tracker first_identified
                fi = t["first_identified"] or fs_date
                rem_str, overdue = rem_date_from(fi, severity)
                status   = "Active"
                row_red  = overdue

            xml_vulns[sec].add(vuln_name.lower())
            groups[gkey]["instances"]  = inst_count
            groups[gkey]["vuln_name"]  = vuln_name
            groups[gkey]["status"]     = status
            groups[gkey]["rem_date"]   = rem_str
            groups[gkey]["overdue"]    = overdue
            groups[gkey]["row_red"]    = row_red
            groups[gkey]["first_seen"] = first_seen

        else:
            # SAST / IaC — group by tool code per result
            for result in results:
                tool_el  = result.find("tool")
                code_raw = tool_el.attrib.get("code", "") if tool_el is not None else ""
                vuln_name = clean_code(code_raw) if code_raw else vuln_cat

                gkey = (sec, vuln_name, vuln_cat, severity)

                t = tracker_lookup(tracker, vuln_name, sec)
                if t is None:
                    rem_str, overdue = rem_date_from(fs_date, severity)
                    status  = "New"
                    row_red = False
                elif t["remediation"].strip().lower() == "fixed":
                    rem_str, overdue = rem_date_from(fs_date, severity)
                    status  = "New"
                    row_red = True
                else:
                    fi = t["first_identified"] or fs_date
                    rem_str, overdue = rem_date_from(fi, severity)
                    status  = "Active"
                    row_red = overdue

                xml_vulns[sec].add(vuln_name.lower())
                groups[gkey]["instances"] += 1
                groups[gkey]["vuln_name"]  = vuln_name
                groups[gkey]["status"]     = status
                groups[gkey]["rem_date"]   = rem_str
                groups[gkey]["overdue"]    = overdue
                groups[gkey]["row_red"]    = row_red
                groups[gkey]["first_seen"] = first_seen

    # ── Build active sections ─────────────────────────────────────────────────
    sev_rank = {s: i for i, s in enumerate(SEV_ORDER)}
    sections = {"SAST": [], "SCA": [], "IaC": []}

    for (sec, key2, vuln_cat, severity), data in groups.items():
        sections[sec].append({
            "vuln_cat":  vuln_cat,
            "vuln_name": data.get("vuln_name", key2),
            "severity":  severity,
            "status":    data["status"],
            "instances": data["instances"],
            "rem_date":  data["rem_date"],
            "overdue":   data["overdue"],
            "row_red":   data["row_red"],
        })

    for sec in sections:
        sections[sec].sort(key=lambda x: sev_rank.get(x["severity"].lower(), 99))

    # ── Build fixed sections (tracker Active but not in XML) ──────────────────
    fixed = {"SAST": [], "SCA": [], "IaC": []}
    for (vname_lower, assess_lower), t in tracker.items():
        if t["remediation"].strip().lower() != "active":
            continue
        # Map assess to section key
        sec = None
        for s in ["SAST", "SCA", "IaC"]:
            if s.lower() in assess_lower or assess_lower in s.lower():
                sec = s
                break
        if sec is None:
            continue
        if vname_lower not in xml_vulns[sec]:
            fixed[sec].append({
                "vuln_name": t["vuln_name"],
                "severity":  t["severity"],
                "status":    "Fixed",
            })

    for sec in fixed:
        fixed[sec].sort(key=lambda x: sev_rank.get(x["severity"].lower(), 99))

    return app_name, scan_date, proj_id, portal, author_name, sections, fixed


# ── HTML helpers ──────────────────────────────────────────────────────────────
def summary_counts(findings):
    c = {s: 0 for s in SEV_ORDER}
    for f in findings:
        c[f["severity"].lower()] = c.get(f["severity"].lower(), 0) + 1
    return c


def build_summary_table(counts):
    def cell(sev):
        v = counts.get(sev, 0)
        return (f'<td style="padding:5px 18px;border:1px solid #bbb;'
                f'text-align:center;font-weight:bold;">{v}</td>')
    return (
        '<table style="border-collapse:collapse;font-size:11px;margin:4px 0 8px;">'
        f'<thead><tr style="background:{TBL_HEADER};">'
        '<th style="padding:5px 18px;border:1px solid #bbb;">Critical</th>'
        '<th style="padding:5px 18px;border:1px solid #bbb;">High</th>'
        '<th style="padding:5px 18px;border:1px solid #bbb;">Medium</th>'
        '<th style="padding:5px 18px;border:1px solid #bbb;">Low</th>'
        '</tr></thead><tbody><tr>'
        + cell("critical") + cell("high") + cell("medium") + cell("low")
        + '</tr></tbody></table><br>'
    )


def build_active_table(findings):
    if not findings:
        return '<p style="font-size:11px;color:#888;">No active findings.</p><br>'
    rows = ""
    for i, f in enumerate(findings, 1):
        if f["row_red"]:
            row_style = 'color:#CC0000;'
        elif f["overdue"]:
            row_style = 'color:#CC0000;'
        else:
            row_style = ''
        rem_color = 'color:#CC0000;font-weight:bold;' if f["overdue"] or f["row_red"] else ''
        rows += (
            f'<tr style="{row_style}">'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{i}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;">{f["vuln_cat"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;">{f["vuln_name"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{f["severity"].title()}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{f["status"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;{rem_color}">{f["rem_date"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;font-weight:bold;">{f["instances"]}</td>'
            f'</tr>'
        )
    return (
        '<p style="margin:10px 0 4px;font-size:11px;text-decoration:underline;">Vulnerabilities Requiring Attention:</p>'
        '<table style="border-collapse:collapse;width:100%;font-size:11px;">'
        f'<thead><tr style="background:{TBL_HEADER};">'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">S.No.</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;">Vulnerability Category</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;">Vulnerability Name</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Severity</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Status</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Remediation Due Date</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Instance Count</th>'
        '</tr></thead>'
        f'<tbody>{rows}</tbody></table><br>'
    )


def build_fixed_table(fixed_findings):
    if not fixed_findings:
        return ""
    rows = ""
    for i, f in enumerate(fixed_findings, 1):
        rows += (
            f'<tr>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{i}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;">{f["vuln_name"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{f["severity"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;color:green;font-weight:bold;">Fixed</td>'
            f'</tr>'
        )
    return (
        '<p style="margin:14px 0 4px;font-size:11px;text-decoration:underline;">Fixed Vulnerabilities:</p>'
        '<table style="border-collapse:collapse;width:100%;font-size:11px;">'
        f'<thead><tr style="background:{TBL_HEADER};">'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">S.No.</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;">Vulnerability Name</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Severity</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Status</th>'
        '</tr></thead>'
        f'<tbody>{rows}</tbody></table><br>'
    )


def build_section(label, findings, fixed_findings):
    counts   = summary_counts(findings)
    s_tbl    = build_summary_table(counts)
    a_tbl    = build_active_table(findings)
    f_tbl    = build_fixed_table(fixed_findings)
    return (
        f'<p style="margin:16px 0 4px;font-size:11px;font-weight:bold;">{label} Summary:</p>'
        f'<p style="margin:2px 0 4px;font-size:11px;text-decoration:underline;">Active Vulnerabilities Count:</p>'
        f'{s_tbl}{a_tbl}{f_tbl}'
    )


# ── Full HTML ─────────────────────────────────────────────────────────────────
def generate_html(app_name, scan_date, portal, author_name, sections, fixed, img_src):
    today      = date.today().strftime("%d-%b-%Y")
    sast_block = build_section("SAST", sections["SAST"], fixed["SAST"])
    sca_block  = build_section("SCA",  sections["SCA"],  fixed["SCA"])
    iac_block  = build_section("IaC",  sections["IaC"],  fixed["IaC"])

    img_block   = (f'<img src="{img_src}" style="max-width:480px;width:100%;height:auto;display:block;margin:8px 0;">'
                   if img_src else
                   '<p style="color:#CC0000;font-size:11px;">&lt;insert-SLA-image-here&gt;</p>')
    author_line = author_name if author_name else "&lt;insert name here&gt;"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Security Email Draft – {app_name}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:Calibri,'Segoe UI',sans-serif;font-size:11px;color:#222;
        background:#eef2f7;padding:20px;}}
  .page{{max-width:920px;margin:0 auto;}}
  .toolbar{{background:#1a3a5c;color:#fff;padding:12px 18px;border-radius:8px 8px 0 0;
            display:flex;align-items:center;justify-content:space-between;}}
  .t-title{{font-size:13px;font-weight:bold;}}
  .t-meta{{font-size:10px;color:#a8c4e0;margin-top:3px;}}
  .copy-btn{{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.4);
             color:#fff;padding:6px 14px;border-radius:4px;cursor:pointer;
             font-size:11px;font-family:inherit;transition:background .2s;}}
  .copy-btn:hover{{background:rgba(255,255,255,.28);}}
  .copy-btn.ok{{background:#2e7d32;border-color:#2e7d32;}}
  .card{{background:#fff;border:1px solid #d0d7de;border-top:none;
         border-radius:0 0 8px 8px;padding:28px 32px 32px;line-height:1.8;}}
  .foot{{margin-top:14px;font-size:10px;color:#999;border-top:1px solid #eee;padding-top:8px;}}
  mark{{background:#FFFF00;}}
</style>
</head>
<body>
<div class="page">
  <div class="toolbar">
    <div>
      <div class="t-title">📧 Email Draft — Security Report: {app_name}</div>
      <div class="t-meta">Generated: {today}</div>
    </div>
    <button class="copy-btn" id="cb" onclick="copyEmail()">📋 Copy Email Body</button>
  </div>
  <div class="card">
    <div id="emailBody">
      <p>Hello Team,</p><br>
      <p>The security assessment of the application <strong>{app_name}</strong> has been completed.
      You can view the detailed vulnerabilities here:&nbsp;
      <a href="{portal}" style="color:#CC0000;font-weight:bold;">{portal}</a></p><br>
      <p><strong>Assessment Summary</strong></p>
      <p>Note, this is a list of all vulnerabilities and findings noted as part of the SAST, SCA, IaC and Container security assessment.</p><br>
      <p><strong>Scan Date:</strong> {scan_date}</p>
      <p><strong>Report Shared Date:</strong> {today}</p><br>

      {sast_block}

      {sca_block}

      <p style="font-size:11px;">In addition, please note:</p>
      <ul style="font-size:11px;margin:4px 0 4px 20px;line-height:1.8;">
        <li>16 vulnerability have been analysed and reported as Informational findings.</li>
        <li>163 packages have been identified with legal risks.</li>
        <li>249 packages are identified as outdated.</li>
      </ul>
      <p style="font-size:11px;">More details are available in the Checkmarx dashboards.</p><br>

      {iac_block}

      <p><strong>Action Required</strong></p>
      <p><strong>Kindly share your remediation plan for the reported vulnerabilities. Please refer to the Corteva remediation policy timelines as mentioned below:</strong></p><br>
      {img_block}<br>

      <p><strong>Please let us know if you have any queries or need assistance from our side regarding the remediation or reported vulnerabilities.</strong></p><br>

      <p><strong>Additional Note:</strong></p>
      <ul style="font-size:11px;margin:4px 0 4px 20px;line-height:1.8;">
        <li>Kindly refer Appsec-Wiki for our processes and guideline documents on how to get access to the tools and navigate through them (e.g. for blackduck access you can refer MyAccessRequest_Blackduck.pdf document)</li>
        <li>It is recommended to take the action as stated in Blackduck_UserGuide.pdf available in the wiki link once you receive this report. In case of any query or requirement or update, you may reach out at dl-appSec.</li>
        <li><mark>Changes to production environment should only be made after proper testing is completed in non-prod ensuring compatibility within our environment.</mark></li>
      </ul><br>

      <p>Regards,<br>{author_line}</p>
    </div>
    <div class="foot">
      Auto-generated by generate_email_draft.py on {today}.
      Open in browser → Copy Email Body → Paste into Outlook.
    </div>
  </div>
</div>
<script>
function copyEmail(){{
  var el=document.getElementById('emailBody'),s=window.getSelection(),r=document.createRange();
  r.selectNodeContents(el);s.removeAllRanges();s.addRange(r);
  document.execCommand('copy');s.removeAllRanges();
  var b=document.getElementById('cb');
  b.textContent='✅ Copied!';b.classList.add('ok');
  setTimeout(function(){{b.textContent='📋 Copy Email Body';b.classList.remove('ok');}},2500);
}}
</script>
</body>
</html>"""


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    xml_path  = find_file("xml")
    if not xml_path:
        print("ERROR: No .xml file found."); sys.exit(1)

    xlsx_path = find_file("xlsx")
    img_src   = find_png()

    tracker   = load_tracker(xlsx_path)
    app_name, scan_date, proj_id, portal, author_name, sections, fixed = parse_report(xml_path, tracker)

    print(f"\n  Application : {app_name}  (Project ID: {proj_id})")
    print(f"  Portal Link : {portal}")
    print(f"  Scan Date   : {scan_date}")
    print(f"  Author      : {author_name}")

    for sec in ["SAST", "SCA", "IaC"]:
        c = summary_counts(sections[sec])
        print(f"\n  {sec} Active — Critical:{c.get('critical',0)}  High:{c.get('high',0)}  Medium:{c.get('medium',0)}  Low:{c.get('low',0)}")
        for r in sections[sec]:
            tag = " 🔴 RE-EMERGED" if r["row_red"] and r["status"]=="New" else (" ⚠️ OVERDUE" if r["overdue"] else "")
            print(f"    [{r['severity'].upper():8s}] {r['vuln_name']}  — {r['instances']} instance(s)  [{r['status']}]{tag}")
        if fixed[sec]:
            print(f"  {sec} Fixed:")
            for f in fixed[sec]:
                print(f"    ✅ {f['vuln_name']}  [{f['severity']}]")

    html     = generate_html(app_name, scan_date, portal, author_name, sections, fixed, img_src)
    d        = os.path.dirname(os.path.abspath(__file__))
    safe     = app_name.replace(" ", "_").replace("/", "-")
    out_path = os.path.join(d, f"email_draft_{safe}.html")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n  ✅ Email draft : email_draft_{safe}.html")
    print(f"  📂 Saved at    : {out_path}\n")


if __name__ == "__main__":
    main()
