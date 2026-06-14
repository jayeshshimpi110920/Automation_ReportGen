"""
Checkmarx Security Report — Email Draft Generator
===================================================
Drop in same folder:
  - Checkmarx XML report   (*.xml)
  - Vulnerability Tracker  (*.xlsx)
  - SLA image              (*.png)   optional

Run:  python generate_email_draft.py
Open generated HTML → Copy Email Body → Paste into Outlook.
"""

import xml.etree.ElementTree as ET
import os, glob, sys, base64
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from collections import defaultdict

# ── Config ────────────────────────────────────────────────────────────────────
BASE_URL        = "https://blackduck.phibred.com:444/srm/projects"
SEV_ORDER       = ["critical", "high", "medium", "low", "info"]
SECTION_MAP     = {
    "component analysis": "SCA",
    "static analysis":    "SAST",
    "iac analysis":       "IaC",
}
REM_SLA         = {"critical": None, "high": 45, "medium": 90, "low": 180, "info": None}
TBL_HEADER      = "#BDD7EE"
FUZZY_THRESHOLD = 0.90
# ─────────────────────────────────────────────────────────────────────────────


# ── File finders ──────────────────────────────────────────────────────────────
def find_file(ext):
    d = os.path.dirname(os.path.abspath(__file__))
    files = glob.glob(os.path.join(d, f"*.{ext}"))
    if not files:
        return None
    files.sort(key=os.path.getmtime, reverse=True)
    print(f"Found .{ext:4s} : {os.path.basename(files[0])}")
    return files[0]


def find_png():
    d = os.path.dirname(os.path.abspath(__file__))
    files = glob.glob(os.path.join(d, "*.png"))
    if not files:
        return None
    files.sort(key=os.path.getmtime, reverse=True)
    with open(files[0], "rb") as f:
        data = base64.b64encode(f.read()).decode()
    return f"data:image/png;base64,{data}"


# ── Date helpers ──────────────────────────────────────────────────────────────
def parse_date(s):
    if s is None:
        return None
    if hasattr(s, 'hour'):
        return s.date()
    if hasattr(s, 'year') and not hasattr(s, 'hour'):
        return s
    s = str(s).strip()
    if not s or s.lower() == 'none':
        return None
    s_clean = s.replace("Z", "").replace("z", "")
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s_clean[:len(fmt) + 4], fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s_clean).date()
    except Exception:
        return None


def calc_rem_date(first_date, severity):
    sev = severity.lower()
    if sev == "critical":
        return "N/A", False
    if sev == "info":
        return "When possible", False
    days = REM_SLA.get(sev)
    if days is None or first_date is None:
        return "N/A", False
    rd      = first_date + timedelta(days=days)
    overdue = rd < date.today()
    return rd.strftime("%d-%b-%Y"), overdue


def clean_code(s):
    return s.replace("_", " ").strip() if s else ""


def get_av(result, key):
    """Get value of <additional-value key="..."> from a result element."""
    key_lower = key.lower()
    for node in result.findall(".//additional-value") + result.findall(".//Additional-value"):
        if node.attrib.get("key", "").strip().lower() == key_lower:
            return (node.text or "").strip()
    return ""


# ── Fuzzy match ───────────────────────────────────────────────────────────────
def fuzzy_match(a, b):
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio() >= FUZZY_THRESHOLD


# ── Normalize SCA package name for matching ───────────────────────────────────
def normalize_pkg(name):
    """Azure.Identity:1.8.1 and Azure.Identity@1.8.1 → azure.identity:1.8.1"""
    return name.lower().strip().replace("@", ":").replace(" ", "")


# ── Load tracker xlsx ─────────────────────────────────────────────────────────
def load_tracker(xlsx_path):
    if xlsx_path is None:
        print("  No tracker xlsx found — all findings treated as New.")
        return []
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed. Run: pip install openpyxl")
        return []

    wb    = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if "tracker" in name.lower():
            sheet = wb[name]; break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("  Tracker is empty."); return []

    header = [str(c).strip() if c else "" for c in rows[0]]
    print(f"  Tracker columns: {[h for h in header if h]}")

    def col(keyword):
        for i, h in enumerate(header):
            if keyword.lower() in h.lower():
                return i
        return None

    ci_vuln   = col("vulnerability name")
    ci_first  = col("first identified")
    ci_sev    = col("vulnerability severity")
    ci_rem    = col("remediation status")
    ci_assess = col("assessment type")

    missing = [n for n, c in [("Vulnerability Name", ci_vuln),
                                ("First Identified Date", ci_first),
                                ("Vulnerability Severity", ci_sev),
                                ("Remediation Status", ci_rem),
                                ("Assessment Type", ci_assess)] if c is None]
    if missing:
        print(f"  WARNING: Columns not found: {missing}")

    tracker_rows = []
    for row in rows[1:]:
        try:
            raw_vuln = row[ci_vuln] if ci_vuln is not None and ci_vuln < len(row) else None
            if not raw_vuln: continue
            vuln_name = str(raw_vuln).strip()
            if not vuln_name or vuln_name.lower() == "none": continue

            first_id   = parse_date(row[ci_first])  if ci_first  is not None and ci_first  < len(row) else None
            severity   = str(row[ci_sev]).strip()    if ci_sev    is not None and ci_sev    < len(row) and row[ci_sev] else ""
            rem        = str(row[ci_rem]).strip()    if ci_rem    is not None and ci_rem    < len(row) and row[ci_rem] else "Active"
            assess_raw = str(row[ci_assess]).strip() if ci_assess is not None and ci_assess < len(row) and row[ci_assess] else ""

            assess = ""
            al = assess_raw.lower()
            if   "sast" in al or "static"        in al: assess = "SAST"
            elif "sca"  in al or "component"      in al: assess = "SCA"
            elif "iac"  in al or "infrastructure" in al: assess = "IaC"

            tracker_rows.append({
                "vuln_name":        vuln_name,
                "vuln_name_lower":  vuln_name.lower().strip(),
                "vuln_name_norm":   normalize_pkg(vuln_name),
                "first_identified": first_id,
                "severity":         severity,
                "remediation":      rem,
                "assessment":       assess,
            })
        except Exception:
            continue

    print(f"  Tracker rows loaded: {len(tracker_rows)}")
    return tracker_rows


# ── Tracker lookup ────────────────────────────────────────────────────────────
def tracker_lookup(tracker_rows, vuln_name, section):
    """
    SCA  → normalized exact match (handles : vs @ separator, case-insensitive)
    SAST/IaC → exact first, then fuzzy 90%+
    Priority: same section > any section
    """
    vl       = vuln_name.lower().strip()
    vn       = normalize_pkg(vuln_name)

    same_exact = None
    same_fuzzy = None
    any_exact  = None
    any_fuzzy  = None

    for row in tracker_rows:
        same_sec = (row["assessment"] == section)

        if section == "SCA":
            is_match = (row["vuln_name_norm"] == vn)
            is_fuzzy = False
        else:
            is_match = (row["vuln_name_lower"] == vl)
            is_fuzzy = fuzzy_match(vuln_name, row["vuln_name"]) if not is_match else False

        if is_match and same_sec  and same_exact is None: same_exact = row
        if is_fuzzy and same_sec  and same_fuzzy is None: same_fuzzy = row
        if is_match and not same_sec and any_exact is None: any_exact = row
        if is_fuzzy and not same_sec and any_fuzzy is None: any_fuzzy = row

    return same_exact or same_fuzzy or any_exact or any_fuzzy


# ── Parse XML ─────────────────────────────────────────────────────────────────
def parse_report(xml_path, tracker_rows):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    raw_date = root.attrib.get("date", "")
    try:
        scan_dt       = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
        scan_date     = scan_dt.strftime("%d-%b-%Y")
        scan_date_obj = scan_dt.date()
    except Exception:
        scan_date     = raw_date[:10] if raw_date else "N/A"
        scan_date_obj = date.today()

    proj       = root.find("project")
    app_name   = proj.attrib.get("name", "N/A") if proj is not None else "N/A"
    proj_id    = proj.attrib.get("id",   "")    if proj is not None else ""
    portal     = f"{BASE_URL}/{proj_id}"         if proj_id else BASE_URL

    author_raw  = root.attrib.get("author", "")
    author_name = ""
    if author_raw:
        local       = author_raw.split("@")[0]
        author_name = local.replace(".", " ").replace("_", " ").title()

    sev_rank        = {s: i for i, s in enumerate(SEV_ORDER)}
    sections        = {"SAST": [], "SCA": [], "IaC": []}
    xml_seen        = {"SAST": set(), "SCA": set(), "IaC": set()}
    sast_iac_groups = {}

    # SCA groups: key = (component_identifier_normalized, severity, vuln_cat)
    # value = { component_identifier (original), vuln_cat, severity, count,
    #           first_seen (for date calc) }
    sca_groups = defaultdict(lambda: {
        "comp_id": "", "vuln_cat": "", "severity": "",
        "confirmed_count": 0, "first_seen": None
    })

    for finding in root.iter("finding"):
        det = finding.attrib.get("detection-method", "").lower()
        sec = SECTION_MAP.get(det)
        if sec is None:
            continue

        first_seen = finding.attrib.get("first-seen", "")
        fs_date    = parse_date(first_seen) or scan_date_obj

        rule_el  = finding.find("rule")
        vuln_cat = clean_code(rule_el.attrib.get("name", "Unknown")) if rule_el is not None else "Unknown"
        results  = finding.findall(".//result")

        # ── SCA ───────────────────────────────────────────────────────────────
        if sec == "SCA":
            for result in results:
                # Only count CONFIRMED results
                state = get_av(result, "Checkmarx One State")
                if state.upper() != "CONFIRMED":
                    continue

                comp_id  = get_av(result, "Component Identifier")
                if not comp_id:
                    continue

                severity = result.attrib.get("severity", "").lower()
                if not severity:
                    severity = finding.attrib.get("severity", "low").lower()

                gkey = (normalize_pkg(comp_id), severity, vuln_cat.lower())

                sca_groups[gkey]["comp_id"]         = comp_id
                sca_groups[gkey]["vuln_cat"]        = vuln_cat
                sca_groups[gkey]["severity"]        = severity
                sca_groups[gkey]["confirmed_count"] += 1
                # Keep earliest first_seen for SLA calc
                if sca_groups[gkey]["first_seen"] is None:
                    sca_groups[gkey]["first_seen"] = fs_date

                xml_seen[sec].add(normalize_pkg(comp_id))

        # ── SAST / IaC ────────────────────────────────────────────────────────
        else:
            severity = finding.attrib.get("severity", "low").lower()
            for result in results:
                tool_el   = result.find("tool")
                code_raw  = tool_el.attrib.get("code", "") if tool_el is not None else ""
                vuln_name = clean_code(code_raw) if code_raw else vuln_cat

                xml_seen[sec].add(vuln_name.lower().strip())

                gkey = (sec, vuln_name.lower().strip(), severity)
                if gkey not in sast_iac_groups:
                    t = tracker_lookup(tracker_rows, vuln_name, sec)
                    if t is None:
                        status           = "New"
                        rem_str, overdue = calc_rem_date(fs_date, severity)
                    elif t["remediation"].strip().lower() == "fixed":
                        status           = "New"
                        rem_str, overdue = calc_rem_date(fs_date, severity)
                    else:
                        status           = "Active"
                        fi_date          = t["first_identified"] or fs_date
                        rem_str, overdue = calc_rem_date(fi_date, severity)

                    sast_iac_groups[gkey] = {
                        "vuln_cat":  vuln_cat,
                        "vuln_name": vuln_name,
                        "severity":  severity,
                        "status":    status,
                        "instances": 0,
                        "rem_date":  rem_str,
                        "overdue":   overdue,
                    }
                sast_iac_groups[gkey]["instances"] += 1

    # ── Flatten SAST/IaC ──────────────────────────────────────────────────────
    for (sec, _, __), data in sast_iac_groups.items():
        sections[sec].append(data)

    # ── Flatten SCA groups ────────────────────────────────────────────────────
    for (comp_norm, severity, _), data in sca_groups.items():
        if data["confirmed_count"] == 0:
            continue

        comp_id  = data["comp_id"]
        vuln_cat = data["vuln_cat"]
        fs_date  = data["first_seen"] or scan_date_obj

        t = tracker_lookup(tracker_rows, comp_id, "SCA")
        if t is None:
            status           = "New"
            rem_str, overdue = calc_rem_date(fs_date, severity)
        elif t["remediation"].strip().lower() == "fixed":
            status           = "New"
            rem_str, overdue = calc_rem_date(fs_date, severity)
        else:
            status           = "Active"
            fi_date          = t["first_identified"] or fs_date
            rem_str, overdue = calc_rem_date(fi_date, severity)

        sections["SCA"].append({
            "vuln_cat":  vuln_cat,
            "vuln_name": comp_id,      # Component Identifier as name
            "severity":  severity,
            "status":    status,
            "instances": data["confirmed_count"],
            "rem_date":  rem_str,
            "overdue":   overdue,
        })

    # Sort all sections by severity
    for sec in sections:
        sections[sec].sort(key=lambda x: sev_rank.get(x["severity"].lower(), 99))

    # ── Fixed table ───────────────────────────────────────────────────────────
    fixed = {"SAST": [], "SCA": [], "IaC": []}

    for row in tracker_rows:
        if row["remediation"].strip().lower() != "active":
            continue
        sec = row["assessment"]
        if sec not in fixed:
            continue

        if sec == "SCA":
            found_in_xml = row["vuln_name_norm"] in xml_seen["SCA"]
        else:
            found_in_xml = any(
                fuzzy_match(row["vuln_name"], seen_name)
                for seen_name in xml_seen[sec]
            )

        if not found_in_xml:
            fixed[sec].append({
                "vuln_name": row["vuln_name"],
                "severity":  row["severity"],
                "status":    "Fixed",
            })

    for sec in fixed:
        fixed[sec].sort(key=lambda x: sev_rank.get(x["severity"].lower(), 99))

    return app_name, scan_date, proj_id, portal, author_name, sections, fixed


# ── Summary counts ────────────────────────────────────────────────────────────
def summary_counts(findings):
    c = {s: 0 for s in SEV_ORDER}
    for f in findings:
        sev = f["severity"].lower()
        if sev in c:
            c[sev] += 1
    return c


# ── HTML: summary count table ─────────────────────────────────────────────────
def build_summary_table(counts):
    def cell(sev):
        v = counts.get(sev, 0)
        return (f'<td style="padding:5px 18px;border:1px solid #bbb;'
                f'text-align:center;font-weight:bold;">{v}</td>')
    return (
        '<table style="border-collapse:collapse;font-size:11px;margin:4px 0 8px;">'
        f'<thead><tr style="background:{TBL_HEADER};">'
        '<th style="padding:5px 18px;border:1px solid #bbb;">Critical</th>'
        '<th style="padding:5px 18px;border:1px solid #bbb;">High</th>'
        '<th style="padding:5px 18px;border:1px solid #bbb;">Medium</th>'
        '<th style="padding:5px 18px;border:1px solid #bbb;">Low</th>'
        '</tr></thead><tbody><tr>'
        + cell("critical") + cell("high") + cell("medium") + cell("low")
        + '</tr></tbody></table><br>'
    )


# ── HTML: active findings table ───────────────────────────────────────────────
def build_active_table(findings, is_sca=False):
    if not findings:
        return '<p style="font-size:11px;color:#888;margin-bottom:8px;">No active findings.</p><br>'

    vuln_col_header = "Component Identified" if is_sca else "Vulnerability Name"
    rows = ""
    for i, f in enumerate(findings, 1):
        date_style = 'color:#CC0000;font-weight:bold;' if f["overdue"] else 'color:#000000;'
        rows += (
            f'<tr>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{i}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;">{f["vuln_cat"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;">{f["vuln_name"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{f["severity"].title()}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{f["status"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;{date_style}">{f["rem_date"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;font-weight:bold;">{f["instances"]}</td>'
            f'</tr>'
        )
    return (
        '<p style="margin:10px 0 4px;font-size:11px;text-decoration:underline;">'
        'Vulnerabilities Requiring Attention:</p>'
        '<table style="border-collapse:collapse;width:100%;font-size:11px;">'
        f'<thead><tr style="background:{TBL_HEADER};">'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">S.No.</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;">Vulnerability Category</th>'
        f'<th style="padding:5px 8px;border:1px solid #bbb;">{vuln_col_header}</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Severity</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Status</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Remediation Due Date</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Instance Count</th>'
        '</tr></thead>'
        f'<tbody>{rows}</tbody></table><br>'
    )


# ── HTML: fixed findings table ────────────────────────────────────────────────
def build_fixed_table(fixed_findings):
    if not fixed_findings:
        return ""

    # Fixed count summary
    sev_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    for f in fixed_findings:
        sev = f["severity"].lower()
        if sev in sev_counts:
            sev_counts[sev] += 1

    def count_cell(sev):
        v = sev_counts.get(sev, 0)
        return (f'<td style="padding:5px 18px;border:1px solid #bbb;'
                f'text-align:center;font-weight:bold;">{v}</td>')

    count_table = (
        '<p style="margin:14px 0 4px;font-size:11px;text-decoration:underline;">'
        'Fixed Vulnerabilities Count:</p>'
        '<table style="border-collapse:collapse;font-size:11px;margin:4px 0 8px;">'
        f'<thead><tr style="background:{TBL_HEADER};">'
        '<th style="padding:5px 18px;border:1px solid #bbb;">Critical</th>'
        '<th style="padding:5px 18px;border:1px solid #bbb;">High</th>'
        '<th style="padding:5px 18px;border:1px solid #bbb;">Medium</th>'
        '<th style="padding:5px 18px;border:1px solid #bbb;">Low</th>'
        '</tr></thead><tbody><tr>'
        + count_cell("critical") + count_cell("high")
        + count_cell("medium")   + count_cell("low")
        + '</tr></tbody></table><br>'
    )

    rows = ""
    for i, f in enumerate(fixed_findings, 1):
        rows += (
            f'<tr>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{i}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;">{f["vuln_name"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{f["severity"]}</td>'
            f'<td style="padding:5px 8px;border:1px solid #ddd;text-align:center;'
            f'color:green;font-weight:bold;">Fixed</td>'
            f'</tr>'
        )

    detail_table = (
        '<p style="margin:10px 0 4px;font-size:11px;text-decoration:underline;">'
        'Fixed Vulnerabilities:</p>'
        '<table style="border-collapse:collapse;width:100%;font-size:11px;">'
        f'<thead><tr style="background:{TBL_HEADER};">'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">S.No.</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;">Vulnerability Name</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Severity</th>'
        '<th style="padding:5px 8px;border:1px solid #bbb;text-align:center;">Status</th>'
        '</tr></thead>'
        f'<tbody>{rows}</tbody></table><br>'
    )

    return count_table + detail_table


# ── HTML: full section block ──────────────────────────────────────────────────
def build_section(label, findings, fixed_findings):
    counts = summary_counts(findings)
    is_sca = (label == "SCA")
    return (
        f'<p style="margin:16px 0 4px;font-size:11px;font-weight:bold;">{label} Summary:</p>'
        f'<p style="margin:2px 0 4px;font-size:11px;text-decoration:underline;">Active Vulnerabilities Count:</p>'
        + build_summary_table(counts)
        + build_active_table(findings, is_sca=is_sca)
        + build_fixed_table(fixed_findings)
    )


# ── Full HTML page ────────────────────────────────────────────────────────────
def generate_html(app_name, scan_date, portal, author_name, sections, fixed, img_src):
    today      = date.today().strftime("%d-%b-%Y")
    sast_block = build_section("SAST", sections["SAST"], fixed["SAST"])
    sca_block  = build_section("SCA",  sections["SCA"],  fixed["SCA"])
    iac_block  = build_section("IaC",  sections["IaC"],  fixed["IaC"])

    img_block   = (
        f'<img src="{img_src}" style="max-width:480px;width:100%;height:auto;display:block;margin:8px 0;">'
        if img_src else
        '<p style="color:#CC0000;font-size:11px;">&lt;insert-SLA-image-here&gt;</p>'
    )
    author_line = author_name if author_name else "&lt;insert name here&gt;"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Security Email Draft – {app_name}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:Calibri,'Segoe UI',sans-serif;font-size:11px;color:#222;
        background:#eef2f7;padding:20px;}}
  .page{{max-width:920px;margin:0 auto;}}
  .toolbar{{background:#1a3a5c;color:#fff;padding:12px 18px;border-radius:8px 8px 0 0;
            display:flex;align-items:center;justify-content:space-between;}}
  .t-title{{font-size:13px;font-weight:bold;}}
  .t-meta{{font-size:10px;color:#a8c4e0;margin-top:3px;}}
  .copy-btn{{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.4);
             color:#fff;padding:6px 14px;border-radius:4px;cursor:pointer;
             font-size:11px;font-family:inherit;transition:background .2s;}}
  .copy-btn:hover{{background:rgba(255,255,255,.28);}}
  .copy-btn.ok{{background:#2e7d32;border-color:#2e7d32;}}
  .card{{background:#fff;border:1px solid #d0d7de;border-top:none;
         border-radius:0 0 8px 8px;padding:28px 32px 32px;line-height:1.8;}}
  .foot{{margin-top:14px;font-size:10px;color:#999;border-top:1px solid #eee;padding-top:8px;}}
  mark{{background:#FFFF00;}}
</style>
</head>
<body>
<div class="page">
  <div class="toolbar">
    <div>
      <div class="t-title">📧 Email Draft — Security Report: {app_name}</div>
      <div class="t-meta">Generated: {today}</div>
    </div>
    <button class="copy-btn" id="cb" onclick="copyEmail()">📋 Copy Email Body</button>
  </div>
  <div class="card">
    <div id="emailBody">
      <p>Hello Team,</p><br>
      <p>The security assessment of the application <strong>{app_name}</strong> has been completed.
      You can view the detailed vulnerabilities here:&nbsp;
      <a href="{portal}" style="color:#CC0000;font-weight:bold;">{portal}</a></p><br>
      <p><strong>Assessment Summary</strong></p>
      <p>Note, this is a list of all vulnerabilities and findings noted as part of the SAST, SCA, IaC and Container security assessment.</p><br>
      <p><strong>Scan Date:</strong> {scan_date}</p>
      <p><strong>Report Shared Date:</strong> {today}</p><br>

      {sast_block}
      {sca_block}

      <p style="font-size:11px;">In addition, please note:</p>
      <ul style="font-size:11px;margin:4px 0 4px 20px;line-height:1.8;">
        <li>16 vulnerability have been analysed and reported as Informational findings.</li>
        <li>163 packages have been identified with legal risks.</li>
        <li>249 packages are identified as outdated.</li>
      </ul>
      <p style="font-size:11px;">More details are available in the Checkmarx dashboards.</p><br>

      {iac_block}

      <p><strong>Action Required</strong></p>
      <p><strong>Kindly share your remediation plan for the reported vulnerabilities. Please refer to the Corteva remediation policy timelines as mentioned below:</strong></p><br>
      {img_block}<br>

      <p><strong>Please let us know if you have any queries or need assistance from our side regarding the remediation or reported vulnerabilities.</strong></p><br>

      <p><strong>Additional Note:</strong></p>
      <ul style="font-size:11px;margin:4px 0 4px 20px;line-height:1.8;">
        <li>Kindly refer Appsec-Wiki for our processes and guideline documents on how to get access to the tools and navigate through them.</li>
        <li>It is recommended to take the action as stated in Blackduck_UserGuide.pdf available in the wiki link once you receive this report. In case of any query or requirement or update, you may reach out at dl-appSec.</li>
        <li><mark>Changes to production environment should only be made after proper testing is completed in non-prod ensuring compatibility within our environment.</mark></li>
      </ul><br>

      <p>Regards,<br>{author_line}</p>
    </div>
    <div class="foot">
      Auto-generated by generate_email_draft.py on {today}.
      Open in browser → Copy Email Body → Paste into Outlook.
    </div>
  </div>
</div>
<script>
function copyEmail(){{
  var el=document.getElementById('emailBody'),s=window.getSelection(),r=document.createRange();
  r.selectNodeContents(el);s.removeAllRanges();s.addRange(r);
  document.execCommand('copy');s.removeAllRanges();
  var b=document.getElementById('cb');
  b.textContent='✅ Copied!';b.classList.add('ok');
  setTimeout(function(){{b.textContent='📋 Copy Email Body';b.classList.remove('ok');}},2500);
}}
</script>
</body>
</html>"""


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    xml_path  = find_file("xml")
    if not xml_path:
        print("ERROR: No .xml file found."); sys.exit(1)

    xlsx_path    = find_file("xlsx")
    img_src      = find_png()
    tracker_rows = load_tracker(xlsx_path)

    app_name, scan_date, proj_id, portal, author_name, sections, fixed = \
        parse_report(xml_path, tracker_rows)

    print(f"\n  Application : {app_name}  (Project ID: {proj_id})")
    print(f"  Portal Link : {portal}")
    print(f"  Scan Date   : {scan_date}")
    print(f"  Author      : {author_name}\n")

    for sec in ["SAST", "SCA", "IaC"]:
        c = summary_counts(sections[sec])
        print(f"  {sec} — Critical:{c.get('critical',0)}  High:{c.get('high',0)}"
              f"  Medium:{c.get('medium',0)}  Low:{c.get('low',0)}")
        for r in sections[sec]:
            flag = " ⚠️ OVERDUE" if r["overdue"] else ""
            print(f"    [{r['severity'].upper():8s}] {r['vuln_name']}"
                  f"  instances:{r['instances']}  [{r['status']}]  rem:{r['rem_date']}{flag}")
        if fixed[sec]:
            print(f"  {sec} Fixed ({len(fixed[sec])}):")
            for f in fixed[sec]:
                print(f"    ✅ {f['vuln_name']}  [{f['severity']}]")
        print()

    html     = generate_html(app_name, scan_date, portal, author_name, sections, fixed, img_src)
    d        = os.path.dirname(os.path.abspath(__file__))
    safe     = app_name.replace(" ", "_").replace("/", "-")
    out_path = os.path.join(d, f"email_draft_{safe}.html")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"  ✅ Email draft : email_draft_{safe}.html")
    print(f"  📂 Saved at    : {out_path}\n")


if __name__ == "__main__":
    main()
